# -*- coding: utf-8 -*-

import os
import json
import numpy as np
import requests
from pathlib import Path
from pypdf import PdfReader
from openpyxl import load_workbook

# ---- настройки ----
OPENAI_BASE = os.environ.get("OPENAI_BASE", "http://26.81.18.206:1234/v1")
EMBED_MODEL = os.environ.get("EMB_MODEL", "text-embedding-bge-m3")

FILES_DIR  = Path("knowledge/pdf")      # туда же кладём и PDF, и XLSX
INDEX_DIR  = Path("knowledge/index")
INDEX_DIR.mkdir(parents=True, exist_ok=True)

CHUNK_SIZE    = 1200
CHUNK_OVERLAP = 200


# ---- чтение PDF ----
def read_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    texts = []
    for page in reader.pages:
        t = page.extract_text() or ""
        if t.strip():
            texts.append(t)
    return "\n".join(texts)


# ---- чтение XLSX ----
def read_xlsx_text(path: Path) -> str:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"=== Лист: {sheet.title} ===")
        for row in sheet.iter_rows(values_only=True):
            vals = [str(v).strip() for v in row if v not in (None, "")]
            if vals:
                lines.append(" | ".join(vals))
        lines.append("")  # разделитель между листами
    wb.close()
    return "\n".join(lines)


# ---- нарезка ----
def chunk_text(text: str, size=CHUNK_SIZE, overlap=CHUNK_OVERLAP):
    text = " ".join(text.split())
    chunks = []
    i = 0
    while i < len(text):
        chunks.append(text[i:i+size])
        i += max(1, size - overlap)
    return chunks


# ---- эмбеддинги ----
def embed_texts(texts):
    """возвращает np.ndarray формы (N, D)"""
    url = f"{OPENAI_BASE}/embeddings"
    out = []
    BATCH = 32
    for i in range(0, len(texts), BATCH):
        batch = texts[i:i+BATCH]
        resp = requests.post(url, json={"model": EMBED_MODEL, "input": batch}, timeout=180)
        resp.raise_for_status()
        data = resp.json()["data"]
        for e in data:
            out.append(e["embedding"])
    arr = np.array(out, dtype="float32")
    norms = np.linalg.norm(arr, axis=1, keepdims=True)
    arr = arr / np.clip(norms, 1e-8, None)
    return arr


# ---- сбор индекса ----
def build_index():
    files = sorted([p for p in FILES_DIR.glob("*") if p.suffix.lower() in (".pdf", ".xlsx")])
    if not files:
        print("Нет PDF/XLSX в knowledge/pdf — положите файлы и запустите снова.")
        return

    records, all_texts = [], []

    for f in files:
        print(f"Читаю: {f.name}")
        if f.suffix.lower() == ".pdf":
            text = read_pdf_text(f)
        elif f.suffix.lower() == ".xlsx":
            text = read_xlsx_text(f)
        else:
            continue

        chunks = chunk_text(text)
        for j, ch in enumerate(chunks):
            records.append({
                "doc": f.name,
                "chunk_id": f"{f.name}__{j}",
                "text": ch
            })
            all_texts.append(ch)

    print(f"Всего чанков: {len(records)}. Считаю эмбеддинги (model={EMBED_MODEL}) …")
    embs = embed_texts(all_texts)

    np.save(INDEX_DIR / "embeddings.npy", embs)
    with open(INDEX_DIR / "records.json", "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    print(f"ОК: {len(records)} чанков, {embs.shape} — индекс в {INDEX_DIR}")


if __name__ == "__main__":
    build_index()
